import psycopg2
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Koneksi ke database PostgreSQL
def connect_db():
    try:
        conn = psycopg2.connect(
            dbname="reservasihotelfahmi", user="postgres", password="1qaz2wsx", host="localhost", port="5432"
        )
        return conn
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

# Fungsi untuk mengekspor data tamu ke Excel
def export_tamu_to_excel():
    conn = connect_db()
    if not conn:
        print("Database connection failed.")
        return

    cursor = conn.cursor()

    # Ambil data tamu dari database
    cursor.execute("SELECT * FROM tamu")
    rows = cursor.fetchall()

    # Membuat workbook dan sheet baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Tamu"

    # Menambahkan header ke sheet Excel
    columns = ['ID Tamu', 'Nama', 'Kontak']  # Sesuaikan dengan kolom di tabel tamu
    for col_num, column_title in enumerate(columns, 1):
        ws[get_column_letter(col_num) + '1'] = column_title

    # Menambahkan data tamu ke Excel
    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            ws[get_column_letter(col_num) + str(row_num)] = cell_value

    # Menyimpan file Excel
    wb.save("data_tamu.xlsx")
    print("Data tamu berhasil diekspor ke data_tamu.xlsx")

# Fungsi untuk mengekspor data kamar ke Excel
def export_kamar_to_excel():
    conn = connect_db()
    if not conn:
        print("Database connection failed.")
        return

    cursor = conn.cursor()

    # Ambil data kamar dari database
    cursor.execute("SELECT * FROM kamar")
    rows = cursor.fetchall()

    # Membuat workbook dan sheet baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Kamar"

    # Menambahkan header ke sheet Excel
    columns = ['ID Kamar', 'Tipe', 'Harga', 'Status']  # Sesuaikan dengan kolom di tabel kamar
    for col_num, column_title in enumerate(columns, 1):
        ws[get_column_letter(col_num) + '1'] = column_title

    # Menambahkan data kamar ke Excel
    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            ws[get_column_letter(col_num) + str(row_num)] = cell_value

    # Menyimpan file Excel
    wb.save("data_kamar.xlsx")
    print("Data kamar berhasil diekspor ke data_kamar.xlsx")

# Fungsi untuk mengekspor data reservasi ke Excel
def export_reservasi_to_excel():
    conn = connect_db()
    if not conn:
        print("Database connection failed.")
        return

    cursor = conn.cursor()

    # Ambil data reservasi dari database
    cursor.execute("SELECT * FROM reservasi")
    rows = cursor.fetchall()

    # Membuat workbook dan sheet baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservasi"

    # Menambahkan header ke sheet Excel
    columns = ['ID Reservasi', 'ID Tamu', 'ID Kamar', 'Check In', 'Check Out', 'Total Harga' ]  # Sesuaikan dengan kolom di tabel reservasi
    for col_num, column_title in enumerate(columns, 1):
        ws[get_column_letter(col_num) + '1'] = column_title

    # Menambahkan data reservasi ke Excel
    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            ws[get_column_letter(col_num) + str(row_num)] = cell_value

    # Menyimpan file Excel
    wb.save("data_reservasi.xlsx")
    print("Data reservasi berhasil diekspor ke data_reservasi.xlsx")
